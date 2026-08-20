"""
Import d'élèves depuis Excel (.xlsx) ou CSV / TSV.

Colonnes attendues :
  matricule;numero_identification;numero_permanent;numero_impot;nom;postnom;prenom;
  date_naissance;sexe;classe;ecole_code;lieu_naissance;adresse;
  nom_pere;postnom_pere;prenom_pere;telephone_pere;email_pere;profession_pere;
  nom_mere;postnom_mere;prenom_mere;telephone_mere;email_mere;profession_mere;
  lien_tuteur;nom_tuteur;telephone_tuteur;email_tuteur

Champs obligatoires : matricule, nom, prenom, date_naissance, sexe, classe
École : ecole_code dans la ligne, sinon ecole_id / ecole_code par défaut.
Classe : nom exact d'une classe déjà créée pour l'école (pas de création auto).
"""
from __future__ import annotations

import csv
import io
import re
from datetime import date, datetime
from typing import Any

from django.db import transaction
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook

from ecoles.models import Classe, Ecole
from .models import Eleve


def _resoudre_classe(ecole, nom_classe: str) -> Classe:
    nom = (nom_classe or '').strip()
    if not nom:
        raise ValueError('classe manquante')
    existante = Classe.objects.filter(ecole=ecole, nom__iexact=nom).first()
    if not existante:
        raise ValueError(
            f'Classe inconnue « {nom} » pour cette école. '
            'Créez-la d\'abord (Paramètres → Structure scolaire / import Classes — national).'
        )
    return existante

REQUIRED = ('matricule', 'nom', 'prenom', 'date_naissance', 'sexe', 'classe')

HEADER_ALIASES = {
    'matricule': 'matricule',
    'mat': 'matricule',
    'numero_identification': 'numero_identification',
    'numero identification': 'numero_identification',
    'n° identification': 'numero_identification',
    'no_identification': 'numero_identification',
    'num_identification': 'numero_identification',
    'nid': 'numero_identification',
    'numero_permanent': 'numero_permanent',
    'numero permanent': 'numero_permanent',
    'n° permanent': 'numero_permanent',
    'no_permanent': 'numero_permanent',
    'num_permanent': 'numero_permanent',
    'npp': 'numero_permanent',
    'numero_impot': 'numero_impot',
    'numero impot': 'numero_impot',
    'n° impot': 'numero_impot',
    'n° impôt': 'numero_impot',
    'numero_impôt': 'numero_impot',
    'nif': 'numero_impot',
    'nom': 'nom',
    'postnom': 'postnom',
    'post-nom': 'postnom',
    'prenom': 'prenom',
    'prénom': 'prenom',
    'date_naissance': 'date_naissance',
    'date naissance': 'date_naissance',
    'date de naissance': 'date_naissance',
    'naissance': 'date_naissance',
    'sexe': 'sexe',
    'genre': 'sexe',
    'classe': 'classe',
    'ecole_code': 'ecole_code',
    'code_ecole': 'ecole_code',
    'code ecole': 'ecole_code',
    'ecole': 'ecole_code',
    'lieu_naissance': 'lieu_naissance',
    'lieu de naissance': 'lieu_naissance',
    'adresse': 'adresse',
    'nom_pere': 'nom_pere',
    'père': 'nom_pere',
    'pere': 'nom_pere',
    'postnom_pere': 'postnom_pere',
    'prenom_pere': 'prenom_pere',
    'telephone_pere': 'telephone_pere',
    'tel_pere': 'telephone_pere',
    'email_pere': 'email_pere',
    'profession_pere': 'profession_pere',
    'nom_mere': 'nom_mere',
    'mère': 'nom_mere',
    'mere': 'nom_mere',
    'postnom_mere': 'postnom_mere',
    'prenom_mere': 'prenom_mere',
    'telephone_mere': 'telephone_mere',
    'tel_mere': 'telephone_mere',
    'email_mere': 'email_mere',
    'profession_mere': 'profession_mere',
    'lien_tuteur': 'lien_tuteur',
    'lien': 'lien_tuteur',
    'nom_tuteur': 'nom_tuteur',
    'tuteur': 'nom_tuteur',
    'telephone_tuteur': 'telephone_tuteur',
    'telephone': 'telephone_tuteur',
    'téléphone_tuteur': 'telephone_tuteur',
    'tel_tuteur': 'telephone_tuteur',
    'email_tuteur': 'email_tuteur',
}

MODELE_HEADERS = [
    'matricule', 'numero_identification', 'numero_permanent', 'numero_impot',
    'nom', 'postnom', 'prenom', 'date_naissance', 'sexe', 'classe',
    'ecole_code', 'lieu_naissance', 'adresse',
    'nom_pere', 'postnom_pere', 'prenom_pere', 'telephone_pere', 'email_pere', 'profession_pere',
    'nom_mere', 'postnom_mere', 'prenom_mere', 'telephone_mere', 'email_mere', 'profession_mere',
    'lien_tuteur', 'nom_tuteur', 'telephone_tuteur', 'email_tuteur',
]

PARENT_TEXT_FIELDS = (
    'nom_pere', 'postnom_pere', 'prenom_pere', 'telephone_pere', 'email_pere', 'profession_pere',
    'nom_mere', 'postnom_mere', 'prenom_mere', 'telephone_mere', 'email_mere', 'profession_mere',
    'nom_tuteur', 'telephone_tuteur', 'email_tuteur',
)


def _parse_lien_tuteur(value: Any) -> str:
    raw = _cell_str(value).lower()
    if not raw:
        return ''
    mapping = {
        'pere': Eleve.LienTuteur.PERE,
        'père': Eleve.LienTuteur.PERE,
        'mere': Eleve.LienTuteur.MERE,
        'mère': Eleve.LienTuteur.MERE,
        'tuteur': Eleve.LienTuteur.TUTEUR,
        'tuteur_legal': Eleve.LienTuteur.TUTEUR,
        'tuteur légal': Eleve.LienTuteur.TUTEUR,
        'oncle_tante': Eleve.LienTuteur.ONCLE_TANTE,
        'oncle': Eleve.LienTuteur.ONCLE_TANTE,
        'tante': Eleve.LienTuteur.ONCLE_TANTE,
        'grand_parent': Eleve.LienTuteur.GRAND_PARENT,
        'grand-parent': Eleve.LienTuteur.GRAND_PARENT,
        'grandparent': Eleve.LienTuteur.GRAND_PARENT,
        'autre': Eleve.LienTuteur.AUTRE,
    }
    if raw in mapping:
        return mapping[raw]
    valid = {c.value for c in Eleve.LienTuteur}
    if raw in valid:
        return raw
    raise ValueError(
        f'Lien tuteur invalide « {value} » '
        '(pere, mere, tuteur, oncle_tante, grand_parent, autre)'
    )


def _normalize_header(value: str) -> str:
    raw = (value or '').strip().lower()
    spaced = re.sub(r'[\s_]+', ' ', raw)
    underscored = spaced.replace(' ', '_')
    return (
        HEADER_ALIASES.get(raw)
        or HEADER_ALIASES.get(spaced)
        or HEADER_ALIASES.get(underscored)
        or underscored
    )


def _detect_dialect(sample: str) -> csv.Dialect:
    try:
        return csv.Sniffer().sniff(sample, delimiters=';,\t|')
    except csv.Error:
        dialect = csv.excel
        dialect.delimiter = ';' if sample.count(';') >= sample.count(',') else ','
        return dialect


def _cell_str(value: Any) -> str:
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _parse_date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = _cell_str(value)
    if not raw:
        raise ValueError('Date de naissance manquante')
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%d.%m.%Y'):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    raise ValueError(f'Date invalide « {raw} » (attendu AAAA-MM-JJ ou JJ/MM/AAAA)')


def _parse_sexe(value: Any) -> str:
    raw = _cell_str(value).upper()
    if raw in ('M', 'H', 'MASCULIN', 'HOMME', 'GARCON', 'GARÇON'):
        return Eleve.Sexe.MASCULIN
    if raw in ('F', 'FEMININ', 'FÉMININ', 'FEMME', 'FILLE'):
        return Eleve.Sexe.FEMININ
    raise ValueError(f'Sexe invalide « {value} » (M ou F)')


def _decode_bytes(data: bytes) -> str:
    for encoding in ('utf-8-sig', 'utf-8', 'cp1252', 'latin-1'):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode('utf-8', errors='replace')


def _rows_from_xlsx(data: bytes) -> list[dict[str, str]]:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers_raw = next(rows_iter)
    except StopIteration as exc:
        raise ValueError('Fichier Excel vide.') from exc

    headers = [_normalize_header(_cell_str(h)) for h in headers_raw]
    if not any(headers):
        raise ValueError('En-tête Excel introuvable.')

    missing = [f for f in REQUIRED if f not in headers]
    if missing:
        raise ValueError(
            'Colonnes obligatoires manquantes : ' + ', '.join(missing)
            + '. Attendu : matricule, nom, prenom, date_naissance, sexe, classe, …'
        )

    rows: list[dict[str, str]] = []
    for i, values in enumerate(rows_iter, start=2):
        if values is None or not any(v not in (None, '') for v in values):
            continue
        row: dict[str, str] = {'_ligne': str(i)}
        for idx, header in enumerate(headers):
            if not header:
                continue
            row[header] = _cell_str(values[idx] if idx < len(values) else '')
        rows.append(row)
    return rows


def _rows_from_csv(contenu: str | bytes) -> list[dict[str, str]]:
    text = _decode_bytes(contenu) if isinstance(contenu, (bytes, bytearray)) else contenu
    text = text.strip()
    if not text:
        raise ValueError('Fichier vide.')

    sample = text[:4096]
    dialect = _detect_dialect(sample)
    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    if not reader.fieldnames:
        raise ValueError('En-tête CSV introuvable.')

    mapping = {_normalize_header(h): h for h in reader.fieldnames if h}
    missing = [f for f in REQUIRED if f not in mapping]
    if missing:
        raise ValueError(
            'Colonnes obligatoires manquantes : ' + ', '.join(missing)
            + '. Attendu : matricule;nom;prenom;date_naissance;sexe;classe;ecole_code;…'
        )

    rows: list[dict[str, str]] = []
    for i, raw in enumerate(reader, start=2):
        if not any((v or '').strip() for v in raw.values()):
            continue
        row = {canon: (raw.get(original) or '').strip() for canon, original in mapping.items()}
        row['_ligne'] = str(i)
        rows.append(row)
    return rows


def lire_lignes(contenu: str | bytes, filename: str = '') -> list[dict[str, str]]:
    """Parse Excel ou CSV et renvoie une liste de dicts normalisés."""
    name = (filename or '').lower()
    if isinstance(contenu, str):
        return _rows_from_csv(contenu)

    data = bytes(contenu)
    if name.endswith(('.xlsx', '.xlsm')) or data[:2] == b'PK':
        try:
            return _rows_from_xlsx(data)
        except Exception as exc:  # noqa: BLE001
            if name.endswith(('.xlsx', '.xlsm')):
                raise ValueError(f'Lecture Excel impossible : {exc}') from exc
    return _rows_from_csv(data)


def importer_eleves(
    contenu: str | bytes,
    *,
    ecole_id: int | None = None,
    ecole_code: str | None = None,
    filename: str = '',
    update_existing: bool = True,
) -> dict[str, Any]:
    """
    Importe les élèves. Retourne un résumé :
    {created, updated, skipped, errors: [{ligne, message}], total}
    """
    rows = lire_lignes(contenu, filename=filename)

    ecole_defaut = None
    if ecole_id:
        ecole_defaut = Ecole.objects.filter(pk=ecole_id).first()
        if not ecole_defaut:
            raise ValueError(f'École introuvable (id={ecole_id}).')
    elif ecole_code:
        ecole_defaut = Ecole.objects.filter(code=ecole_code.strip()).first()
        if not ecole_defaut:
            raise ValueError(f'École introuvable (code={ecole_code}).')

    ecole_cache: dict[str, Ecole] = {}
    created = updated = skipped = 0
    errors: list[dict[str, str]] = []

    with transaction.atomic():
        for row in rows:
            ligne = row.get('_ligne', '?')
            try:
                matricule = row.get('matricule', '').strip()
                if not matricule:
                    raise ValueError('Matricule manquant')

                code = (row.get('ecole_code') or '').strip()
                ecole = ecole_defaut
                if code:
                    if code not in ecole_cache:
                        ecole_cache[code] = Ecole.objects.filter(code=code).first()
                    ecole = ecole_cache[code]
                    if not ecole:
                        raise ValueError(f'École inconnue (code={code})')
                if not ecole:
                    raise ValueError('École non précisée (colonne ecole_code ou école par défaut)')

                classe = _resoudre_classe(ecole, row.get('classe', ''))
                defaults = {
                    'numero_identification': (row.get('numero_identification') or '').strip() or None,
                    'numero_permanent': (row.get('numero_permanent') or '').strip() or None,
                    'numero_impot': (row.get('numero_impot') or '').strip() or None,
                    'nom': row['nom'].strip(),
                    'postnom': row.get('postnom', '').strip(),
                    'prenom': row['prenom'].strip(),
                    'date_naissance': _parse_date(row['date_naissance']),
                    'lieu_naissance': row.get('lieu_naissance', '').strip(),
                    'sexe': _parse_sexe(row['sexe']),
                    'ecole': ecole,
                    'classe': classe,
                    'adresse': row.get('adresse', '').strip(),
                    'lien_tuteur': _parse_lien_tuteur(row.get('lien_tuteur', '')),
                }
                for key in PARENT_TEXT_FIELDS:
                    defaults[key] = row.get(key, '').strip()
                for key in ('nom', 'prenom'):
                    if not defaults[key]:
                        raise ValueError(f'{key} manquant')

                existing = Eleve.objects.filter(matricule=matricule).first()
                if existing:
                    if not update_existing:
                        skipped += 1
                        continue
                    for k, v in defaults.items():
                        setattr(existing, k, v)
                    existing.save()
                    updated += 1
                else:
                    Eleve.objects.create(matricule=matricule, **defaults)
                    created += 1
            except Exception as exc:  # noqa: BLE001
                errors.append({'ligne': ligne, 'message': str(exc)})

    return {
        'created': created,
        'updated': updated,
        'skipped': skipped,
        'errors': errors[:50],
        'errors_count': len(errors),
        'total': len(rows),
    }


def _slug_fichier(texte: str, max_len: int = 60) -> str:
    """Nom de fichier sûr (sans accents problématiques / caractères réservés)."""
    import unicodedata
    raw = (texte or '').strip()
    raw = unicodedata.normalize('NFKD', raw)
    raw = ''.join(c for c in raw if not unicodedata.combining(c))
    raw = re.sub(r'[<>:"/\\|?*\x00-\x1f]+', '', raw)
    raw = re.sub(r'\s+', '_', raw)
    raw = re.sub(r'_+', '_', raw).strip('._')
    return (raw or 'Ecole')[:max_len]


def _nom_fichier_modele_eleves(ecole: Ecole | None) -> str:
    if ecole:
        nom = _slug_fichier(ecole.nom or 'Ecole')
        code = _slug_fichier((ecole.code or str(ecole.id)).strip(), max_len=24)
        return f'Import_eleves_{nom}_{code}.xlsx'
    return 'Import_eleves_modele.xlsx'


def generer_modele_xlsx(ecole_id: int | None = None) -> tuple[bytes, str]:
    """
    Génère le modèle Excel d'import élèves (mise en forme professionnelle).
    Retourne (contenu_bytes, nom_fichier).
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Protection, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from datetime import date as date_cls

    ecole = None
    if ecole_id:
        ecole = Ecole.objects.filter(pk=ecole_id).first()

    classes_noms: list[str] = []
    if ecole:
        classes_noms = list(
            Classe.objects.filter(ecole=ecole, active=True)
            .order_by('nom')
            .values_list('nom', flat=True)
        )
    if not classes_noms:
        classes_noms = ['6ème Primaire', '5ème Primaire']

    ecole_codes: list[str] = []
    if ecole and (ecole.code or '').strip():
        ecole_codes = [(ecole.code or '').strip()]
    else:
        ecole_codes = ['7-136755']

    sexes = ['M', 'F']
    liens = [c.value for c in Eleve.LienTuteur]
    nom_fichier = _nom_fichier_modele_eleves(ecole)

    # Palette RDC / Educ_RDC
    BLEU_NUIT = '062849'
    BLEU = '007FFF'
    BLEU_PROFOND = '0A3D7A'
    JAUNE = 'FCD116'
    ROUGE = 'CE1126'
    GRIS_LIGNE = 'D6DEE8'
    GRIS_FOND = 'F4F7FB'
    BLANC = 'FFFFFF'
    VERT_DOUX = 'E8F5EE'

    font_title = Font(name='Calibri', bold=True, size=18, color=BLANC)
    font_sub = Font(name='Calibri', bold=True, size=12, color=BLEU_NUIT)
    font_body = Font(name='Calibri', size=11, color='1A2332')
    font_muted = Font(name='Calibri', size=10, color='5A6B7D', italic=True)
    font_header = Font(name='Calibri', bold=True, size=10, color=BLANC)
    font_chip = Font(name='Calibri', bold=True, size=10, color=BLEU_NUIT)

    fill_banner = PatternFill('solid', fgColor=BLEU_PROFOND)
    fill_accent = PatternFill('solid', fgColor=JAUNE)
    fill_header = PatternFill('solid', fgColor=BLEU_PROFOND)
    fill_header_req = PatternFill('solid', fgColor='0B5CAD')
    fill_header_list = PatternFill('solid', fgColor='1565C0')
    fill_zebra = PatternFill('solid', fgColor=GRIS_FOND)
    fill_lock = PatternFill('solid', fgColor='E8EEF6')
    fill_ok = PatternFill('solid', fgColor=VERT_DOUX)
    fill_note = PatternFill('solid', fgColor='FFF8E1')

    thin = Border(
        left=Side(style='thin', color=GRIS_LIGNE),
        right=Side(style='thin', color=GRIS_LIGNE),
        top=Side(style='thin', color=GRIS_LIGNE),
        bottom=Side(style='thin', color=GRIS_LIGNE),
    )
    align_c = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_l = Alignment(horizontal='left', vertical='center', wrap_text=True)

    REQUIRED_COLS = {'matricule', 'nom', 'prenom', 'date_naissance', 'sexe', 'classe'}
    LIST_COLS = {'sexe', 'classe', 'ecole_code', 'lien_tuteur'}

    wb = Workbook()

    # ═══════════ Instructions (couverture) ═══════════
    info = wb.active
    info.title = 'Instructions'
    info.sheet_view.showGridLines = False
    info.sheet_properties.tabColor = BLEU_PROFOND

    info.merge_cells('A1:D1')
    info['A1'] = 'Educ_RDC — Modèle d’import des élèves'
    info['A1'].font = font_title
    info['A1'].fill = fill_banner
    info['A1'].alignment = Alignment(horizontal='left', vertical='center')
    info.row_dimensions[1].height = 36
    for col in ('B', 'C', 'D'):
        info[f'{col}1'].fill = fill_banner

    info.merge_cells('A2:D2')
    info['A2'] = ''
    info['A2'].fill = fill_accent
    info.row_dimensions[2].height = 6
    for col in ('B', 'C', 'D'):
        info[f'{col}2'].fill = fill_accent

    ecole_lib = (
        f'{ecole.nom}  ·  Code {ecole.code}' if ecole and ecole.code
        else (ecole.nom if ecole else 'Modèle générique (sélectionnez une école à l’import)')
    )
    info.merge_cells('A3:D3')
    info['A3'] = ecole_lib
    info['A3'].font = font_sub
    info['A3'].alignment = align_l
    info.row_dimensions[3].height = 22

    info.merge_cells('A4:D4')
    info['A4'] = (
        f'Document généré le {date_cls.today().strftime("%d/%m/%Y")} — '
        'Remplir uniquement la feuille « Eleves ». '
        'Les listes (Sexe, Classes, Ecoles, Liens) sont protégées.'
    )
    info['A4'].font = font_muted
    info['A4'].alignment = align_l

    info['A6'] = 'Guide des colonnes'
    info['A6'].font = Font(name='Calibri', bold=True, size=13, color=BLEU_NUIT)
    info.merge_cells('A6:D6')

    headers_guide = ['Champ', 'Obligatoire', 'Type', 'Description']
    for i, h in enumerate(headers_guide, start=1):
        cell = info.cell(7, i, h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_c
        cell.border = thin
    info.row_dimensions[7].height = 20

    guide_rows = [
        ('matricule', 'Oui', 'Texte', 'Format AAAA-0001 (ex. 2026-0001)'),
        ('numero_identification', 'Non', 'Texte', 'Souvent auto — code école + n° d’ordre'),
        ('numero_permanent', 'Non', 'Texte', 'Unique si renseigné'),
        ('numero_impot', 'Non', 'Texte', 'Unique si renseigné'),
        ('nom', 'Oui', 'Texte', 'Nom de famille'),
        ('postnom', 'Non', 'Texte', 'Postnom'),
        ('prenom', 'Oui', 'Texte', 'Prénom'),
        ('date_naissance', 'Oui', 'Date', 'AAAA-MM-JJ'),
        ('sexe', 'Oui', 'Liste', 'Choisir M ou F'),
        ('classe', 'Oui', 'Liste', 'Choisir une classe existante de l’école'),
        ('ecole_code', 'Non*', 'Liste', 'Code école (*sinon école choisie à l’import)'),
        ('lieu_naissance', 'Non', 'Texte', 'Ville / territoire'),
        ('adresse', 'Non', 'Texte', 'Adresse de résidence'),
        ('nom_pere … profession_pere', 'Non', 'Texte', 'Identité et contacts du père'),
        ('nom_mere … profession_mere', 'Non', 'Texte', 'Identité et contacts de la mère'),
        ('lien_tuteur', 'Non', 'Liste', 'Lien de parenté du responsable'),
        ('nom_tuteur / téléphone / e-mail', 'Non', 'Texte', 'Coordonnées du tuteur'),
    ]
    for r_idx, row in enumerate(guide_rows, start=8):
        for c_idx, val in enumerate(row, start=1):
            cell = info.cell(r_idx, c_idx, val)
            cell.font = font_body
            cell.alignment = align_l if c_idx != 2 else align_c
            cell.border = thin
            if r_idx % 2 == 0:
                cell.fill = fill_zebra
            if c_idx == 2 and val == 'Oui':
                cell.fill = fill_ok
                cell.font = font_chip

    note_row = 8 + len(guide_rows) + 1
    info.merge_cells(f'A{note_row}:D{note_row}')
    info[f'A{note_row}'] = (
        'Ordre recommandé : 1) Créer / importer les classes  →  2) Importer les élèves  →  '
        '3) Créer les comptes enseignants. Ne pas modifier les feuilles de listes.'
    )
    info[f'A{note_row}'].font = font_muted
    info[f'A{note_row}'].fill = fill_note
    info[f'A{note_row}'].alignment = align_l
    info.row_dimensions[note_row].height = 32

    info.column_dimensions['A'].width = 32
    info.column_dimensions['B'].width = 12
    info.column_dimensions['C'].width = 10
    info.column_dimensions['D'].width = 52
    info.freeze_panes = 'A8'
    info.print_title_rows = '1:2'

    def _feuille_liste(titre: str, entete: str, valeurs: list[str], note: str, tab_color: str) -> str:
        ws = wb.create_sheet(titre)
        ws.sheet_properties.tabColor = tab_color
        ws.sheet_view.showGridLines = False

        ws.merge_cells('A1:B1')
        ws['A1'] = f'Référentiel — {entete}'
        ws['A1'].font = Font(name='Calibri', bold=True, size=12, color=BLANC)
        ws['A1'].fill = fill_banner
        ws['A1'].alignment = align_l
        ws['B1'].fill = fill_banner
        ws.row_dimensions[1].height = 24

        ws['A2'] = 'Valeur'
        ws['B2'] = 'Statut'
        for col in (1, 2):
            cell = ws.cell(2, col)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_c
            cell.border = thin
            cell.protection = Protection(locked=True)

        for i, v in enumerate(valeurs, start=3):
            c1 = ws.cell(i, 1, v)
            c2 = ws.cell(i, 2, 'Protégé')
            for c in (c1, c2):
                c.font = font_body
                c.border = thin
                c.protection = Protection(locked=True)
                c.fill = fill_lock
            c1.alignment = align_l
            c2.alignment = align_c
            c2.font = font_chip

        note_r = 3 + len(valeurs) + 1
        ws.merge_cells(f'A{note_r}:B{note_r}')
        ws[f'A{note_r}'] = note
        ws[f'A{note_r}'].font = font_muted
        ws[f'A{note_r}'].fill = fill_note

        ws.column_dimensions['A'].width = max(22, min(48, max((len(str(v)) for v in valeurs), default=12) + 4))
        ws.column_dimensions['B'].width = 12
        ws.protection.sheet = True
        ws.protection.enable()
        ws.protection.password = 'EducRDC'
        ws.freeze_panes = 'A3'
        return titre

    sheet_sexe = _feuille_liste(
        'Sexe', 'sexe', sexes,
        'Feuille protégée — ne pas modifier ces valeurs.',
        ROUGE,
    )
    sheet_liens = _feuille_liste(
        'Liens', 'lien_tuteur', liens,
        'Feuille protégée — ne pas modifier ces valeurs.',
        JAUNE,
    )
    sheet_classes = _feuille_liste(
        'Classes', 'classe', classes_noms,
        'Feuille protégée — classes actives de l’école. Ne pas modifier.',
        BLEU,
    )
    sheet_ecoles = _feuille_liste(
        'Ecoles', 'ecole_code', ecole_codes,
        'Feuille protégée — code école autorisé. Ne pas modifier.',
        BLEU_NUIT,
    )

    # ═══════════ Eleves (saisie) ═══════════
    ws = wb.create_sheet('Eleves', 1)
    ws.sheet_properties.tabColor = '1B7A4E'
    ws.sheet_view.showGridLines = False

    # Bandeau info au-dessus des en-têtes ? — l’import lit la 1re ligne = headers.
    # On met donc le bandeau en commentaire de cellule A1 + style pro des headers.
    ws.append(MODELE_HEADERS)
    for col_idx, name in enumerate(MODELE_HEADERS, start=1):
        cell = ws.cell(1, col_idx)
        cell.font = font_header
        if name in REQUIRED_COLS and name in LIST_COLS:
            cell.fill = fill_header_list
        elif name in REQUIRED_COLS:
            cell.fill = fill_header_req
        elif name in LIST_COLS:
            cell.fill = fill_header_list
        else:
            cell.fill = fill_header
        cell.alignment = align_c
        cell.border = thin
        # Largeurs ciblées
        widths = {
            'matricule': 16, 'numero_identification': 18, 'numero_permanent': 16, 'numero_impot': 14,
            'nom': 14, 'postnom': 14, 'prenom': 14, 'date_naissance': 14, 'sexe': 8,
            'classe': 18, 'ecole_code': 12, 'lieu_naissance': 16, 'adresse': 28,
            'email_pere': 26, 'email_mere': 26, 'email_tuteur': 26,
            'telephone_pere': 16, 'telephone_mere': 16, 'telephone_tuteur': 16,
            'lien_tuteur': 14, 'nom_tuteur': 20, 'profession_pere': 16, 'profession_mere': 16,
        }
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(name, 14)

    ws.row_dimensions[1].height = 32
    ws.auto_filter.ref = f'A1:{get_column_letter(len(MODELE_HEADERS))}1'
    ws.freeze_panes = 'A2'
    ws.print_title_rows = '1:1'

    demo_classe = classes_noms[0]
    demo_ecole = ecole_codes[0]
    demos = [
        [
            'ELV-2026-DEMO-001', 'ID-KIN-001', 'NP-2026-0001', '',
            'KABONGO', 'MUTOMBO', 'Jean', '2015-03-12', 'M',
            demo_classe, demo_ecole, 'Kinshasa', 'Av. de la Libération',
            'KABONGO', 'ILUNGA', 'Pierre', '+243810000010', 'pierre.kabongo@email.cd', 'Commerçant',
            'MWAMBA', 'KABONGO', 'Jeanne', '+243810000011', 'jeanne.mwamba@email.cd', 'Enseignante',
            'mere', 'MWAMBA Jeanne', '+243810000011', 'jeanne.mwamba@email.cd',
        ],
        [
            'ELV-2026-DEMO-002', 'ID-KIN-002', 'NP-2026-0002', '',
            'MUKENDI', '', 'Marie', '2016-07-21', 'F',
            classes_noms[1] if len(classes_noms) > 1 else demo_classe, demo_ecole, 'Kinshasa', 'C/Gombe',
            'MUKENDI', '', 'Joseph', '+243810000020', '', 'Chauffeur',
            'KALALA', '', 'Sophie', '+243810000021', '', '',
            'pere', 'MUKENDI Joseph', '+243810000020', '',
        ],
    ]
    for r_i, demo in enumerate(demos):
        ws.append(demo)
        row_num = 2 + r_i
        for col_idx in range(1, len(MODELE_HEADERS) + 1):
            cell = ws.cell(row_num, col_idx)
            cell.font = font_body
            cell.border = thin
            cell.alignment = align_l
            if r_i % 2 == 1:
                cell.fill = fill_zebra
            cell.protection = Protection(locked=False)

    # Zone de saisie étendue (lignes vides prêtes)
    for row_num in range(4, 201):
        for col_idx in range(1, len(MODELE_HEADERS) + 1):
            cell = ws.cell(row_num, col_idx)
            cell.border = thin
            cell.protection = Protection(locked=False)
            if row_num % 2 == 0:
                cell.fill = fill_zebra

    # Zebra pour lignes encore vides via règle (optionnel, déjà appliqué)

    def _col(name: str) -> str:
        return get_column_letter(MODELE_HEADERS.index(name) + 1)

    def _ajouter_validation(feuille_liste: str, nb: int, col_name: str, obligatoire: bool = False):
        # Valeurs en A3… (ligne 1 titre, ligne 2 en-tête)
        debut = 3
        fin = debut + max(nb, 1) - 1
        formule = f"'{feuille_liste}'!$A${debut}:$A${fin}"
        dv = DataValidation(
            type='list',
            formula1=formule,
            allow_blank=not obligatoire,
            showDropDown=False,
            showErrorMessage=True,
            showInputMessage=True,
            errorTitle='Valeur non autorisée',
            error='Choisissez une valeur dans la liste déroulante.',
            promptTitle=col_name,
            prompt='Sélectionnez une valeur dans la liste (référentiel protégé).',
        )
        dv.errorStyle = 'stop'
        lettre = _col(col_name)
        dv.add(f'{lettre}2:{lettre}2000')
        ws.add_data_validation(dv)

    _ajouter_validation(sheet_sexe, len(sexes), 'sexe', obligatoire=True)
    _ajouter_validation(sheet_classes, len(classes_noms), 'classe', obligatoire=True)
    _ajouter_validation(sheet_ecoles, len(ecole_codes), 'ecole_code', obligatoire=False)
    _ajouter_validation(sheet_liens, len(liens), 'lien_tuteur', obligatoire=False)

    # Propriétés document
    wb.properties.title = f'Import élèves — {ecole.nom if ecole else "Modèle"}'
    wb.properties.creator = 'Educ_RDC'
    wb.properties.description = (
        f'Modèle d’import élèves ({ecole_lib}). Listes déroulantes protégées.'
    )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), nom_fichier


def reponse_modele_xlsx(ecole_id: int | None = None) -> HttpResponse:
    contenu, nom_fichier = generer_modele_xlsx(ecole_id=ecole_id)
    # RFC 5987 pour accents dans le nom
    from urllib.parse import quote
    disposition = (
        f"attachment; filename=\"{nom_fichier}\"; "
        f"filename*=UTF-8''{quote(nom_fichier)}"
    )
    response = HttpResponse(
        contenu,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = disposition
    return response
