"""
Import des classes scolaires depuis Excel (.xlsx) ou CSV.

Colonnes :
  nom;code;ecole_code;active

Obligatoires : nom
École : ecole_code dans la ligne, sinon ecole_id / ecole_code par défaut.
"""
from __future__ import annotations

import csv
import io
import re
from typing import Any

from django.db import transaction
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook

from ecoles.models import Classe, Ecole

REQUIRED = ('nom',)

HEADER_ALIASES = {
    'nom': 'nom',
    'classe': 'nom',
    'nom_classe': 'nom',
    'nom classe': 'nom',
    'code': 'code',
    'code_classe': 'code',
    'ecole_code': 'ecole_code',
    'code_ecole': 'ecole_code',
    'code ecole': 'ecole_code',
    'ecole': 'ecole_code',
    'active': 'active',
    'actif': 'active',
    'activee': 'active',
}

MODELE_HEADERS = ['nom', 'code', 'ecole_code', 'active']


def _normalize_header(value: str) -> str:
    raw = (value or '').strip().lower()
    spaced = re.sub(r'[\s_]+', ' ', raw)
    underscored = spaced.replace(' ', '_')
    return (
        HEADER_ALIASES.get(raw)
        or HEADER_ALIASES.get(spaced)
        or HEADER_ALIASES.get(underscored)
        or underscored
    )


def _cell_str(value: Any) -> str:
    if value is None:
        return ''
    return str(value).strip()


def _parse_bool(value: Any, default: bool = True) -> bool:
    if value is None or value == '':
        return default
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in ('1', 'true', 'oui', 'o', 'yes', 'y', 'actif', 'active'):
        return True
    if text in ('0', 'false', 'non', 'n', 'no', 'inactif', 'inactive'):
        return False
    return default


def _rows_from_xlsx(content: bytes) -> list[dict[str, str]]:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise ValueError('Fichier Excel vide.')
    headers = [_normalize_header(_cell_str(h)) for h in header_row]
    missing = [f for f in REQUIRED if f not in headers]
    if missing:
        raise ValueError('Colonnes obligatoires manquantes : ' + ', '.join(missing))
    out = []
    for i, row in enumerate(rows_iter, start=2):
        if not row or all(v is None or str(v).strip() == '' for v in row):
            continue
        data = {headers[j]: _cell_str(row[j]) for j in range(min(len(headers), len(row)))}
        data['_ligne'] = str(i)
        out.append(data)
    return out


def _rows_from_csv(content: bytes, filename: str = '') -> list[dict[str, str]]:
    text = content.decode('utf-8-sig')
    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=';,\t|')
    except csv.Error:
        dialect = csv.excel
        dialect.delimiter = ';' if sample.count(';') >= sample.count(',') else ','
    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    if not reader.fieldnames:
        raise ValueError('En-têtes CSV introuvables.')
    mapping = {_normalize_header(h): h for h in reader.fieldnames if h}
    missing = [f for f in REQUIRED if f not in mapping]
    if missing:
        raise ValueError('Colonnes obligatoires manquantes : ' + ', '.join(missing))
    out = []
    for i, row in enumerate(reader, start=2):
        data = {canon: (row.get(src) or '').strip() for canon, src in mapping.items()}
        if not any(data.values()):
            continue
        data['_ligne'] = str(i)
        out.append(data)
    return out


def importer_classes(
    content: bytes,
    *,
    ecole_id: int | None = None,
    ecole_code: str | None = None,
    filename: str = '',
    update_existing: bool = True,
) -> dict[str, Any]:
    name = (filename or '').lower()
    if name.endswith(('.xlsx', '.xlsm')):
        rows = _rows_from_xlsx(content)
    else:
        rows = _rows_from_csv(content, filename)

    ecole_defaut = None
    if ecole_id:
        ecole_defaut = Ecole.objects.filter(pk=ecole_id).first()
    elif ecole_code:
        ecole_defaut = Ecole.objects.filter(code__iexact=ecole_code.strip()).first()

    ecole_cache: dict[str, Ecole | None] = {}
    created = updated = skipped = 0
    errors: list[dict[str, str]] = []

    with transaction.atomic():
        for row in rows:
            ligne = row.get('_ligne', '?')
            try:
                nom = row.get('nom', '').strip()
                if not nom:
                    raise ValueError('Nom de classe manquant')

                code_ecole = (row.get('ecole_code') or '').strip()
                ecole = ecole_defaut
                if code_ecole:
                    if code_ecole not in ecole_cache:
                        ecole_cache[code_ecole] = Ecole.objects.filter(code__iexact=code_ecole).first()
                    ecole = ecole_cache[code_ecole]
                    if not ecole:
                        raise ValueError(f'École inconnue (code={code_ecole})')
                if not ecole:
                    raise ValueError('École non précisée (colonne ecole_code ou école par défaut)')

                defaults = {
                    'code': (row.get('code') or '').strip(),
                    'active': _parse_bool(row.get('active'), True),
                }
                existing = Classe.objects.filter(ecole=ecole, nom__iexact=nom).first()
                if existing:
                    if not update_existing:
                        skipped += 1
                        continue
                    for k, v in defaults.items():
                        setattr(existing, k, v)
                    existing.nom = nom
                    existing.save()
                    updated += 1
                else:
                    Classe.objects.create(ecole=ecole, nom=nom, **defaults)
                    created += 1
            except Exception as exc:  # noqa: BLE001
                errors.append({'ligne': ligne, 'message': str(exc)})

    return {
        'created': created,
        'updated': updated,
        'skipped': skipped,
        'errors': errors[:50],
        'errors_count': len(errors),
        'total': len(rows),
        'ecole_id': ecole_defaut.id if ecole_defaut else None,
        'ecole_nom': ecole_defaut.nom if ecole_defaut else None,
    }


def generer_modele_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Classes'
    ws.append(MODELE_HEADERS)
    ws.append(['6ème Primaire', '6P', '7-136755', 'oui'])
    ws.append(['5ème Primaire', '5P', '7-136755', 'oui'])
    ws.append(['4ème Primaire A', '4P-A', '7-136755', 'oui'])

    info = wb.create_sheet('Instructions', 0)
    info.append(['Champ', 'Description'])
    info.append(['nom', 'Obligatoire — nom exact de la classe (ex: 6ème Primaire)'])
    info.append(['code', 'Optionnel — code court interne'])
    info.append(['ecole_code', 'Code de l\'école (sinon école choisie à l\'import)'])
    info.append(['active', 'oui/non — défaut: oui'])
    info.append(['', ''])
    info.append(['Ordre recommandé', '1) Classes  2) Élèves  3) Comptes enseignants'])
    info.append(['Note', 'Le nom de classe des élèves doit correspondre exactement à ce fichier.'])

    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max(len(str(c.value or '')) for c in col)
            sheet.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 14), 48)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def reponse_modele_xlsx() -> HttpResponse:
    response = HttpResponse(
        generer_modele_xlsx(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="modele_import_classes.xlsx"'
    return response
