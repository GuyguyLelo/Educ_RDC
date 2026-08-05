"""
Import du personnel scolaire depuis Excel (.xlsx) ou CSV.

Colonnes attendues :
  matricule;nom;postnom;prenom;sexe;fonction;telephone;email;
  date_naissance;date_prise_service

Obligatoires : nom, prenom, sexe
Fonction : directeur, directeur_etudes, enseignant, secretaire,
           comptable, surveillant, prefet, autre (défaut : enseignant)
"""
from __future__ import annotations

import csv
import io
import re
from datetime import date, datetime
from typing import Any

from django.db import transaction
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook

from ecoles.models import Ecole, PersonnelEcole

REQUIRED = ('nom', 'prenom', 'sexe')

HEADER_ALIASES = {
    'matricule': 'matricule',
    'mat': 'matricule',
    'nom': 'nom',
    'postnom': 'postnom',
    'post-nom': 'postnom',
    'prenom': 'prenom',
    'prénom': 'prenom',
    'sexe': 'sexe',
    'genre': 'sexe',
    'fonction': 'fonction',
    'poste': 'fonction',
    'titre': 'fonction',
    'telephone': 'telephone',
    'téléphone': 'telephone',
    'tel': 'telephone',
    'email': 'email',
    'mail': 'email',
    'date_naissance': 'date_naissance',
    'date naissance': 'date_naissance',
    'naissance': 'date_naissance',
    'date_prise_service': 'date_prise_service',
    'date prise service': 'date_prise_service',
    'prise_service': 'date_prise_service',
    'prise de service': 'date_prise_service',
}

FONCTION_ALIASES = {
    'directeur': PersonnelEcole.Fonction.DIRECTEUR,
    'directrice': PersonnelEcole.Fonction.DIRECTEUR,
    'directeur / directrice': PersonnelEcole.Fonction.DIRECTEUR,
    'directeur_etudes': PersonnelEcole.Fonction.DIRECTEUR_ETUDES,
    'directeur des etudes': PersonnelEcole.Fonction.DIRECTEUR_ETUDES,
    'directeur des études': PersonnelEcole.Fonction.DIRECTEUR_ETUDES,
    'enseignant': PersonnelEcole.Fonction.ENSEIGNANT,
    'enseignante': PersonnelEcole.Fonction.ENSEIGNANT,
    'professeur': PersonnelEcole.Fonction.ENSEIGNANT,
    'secretaire': PersonnelEcole.Fonction.SECRETAIRE,
    'secrétaire': PersonnelEcole.Fonction.SECRETAIRE,
    'comptable': PersonnelEcole.Fonction.COMPTABLE,
    'surveillant': PersonnelEcole.Fonction.SURVEILLANT,
    'surveillante': PersonnelEcole.Fonction.SURVEILLANT,
    'prefet': PersonnelEcole.Fonction.PREFET,
    'préfet': PersonnelEcole.Fonction.PREFET,
    'prefete': PersonnelEcole.Fonction.PREFET,
    'préfète': PersonnelEcole.Fonction.PREFET,
    'autre': PersonnelEcole.Fonction.AUTRE,
}

MODELE_HEADERS = [
    'matricule', 'nom', 'postnom', 'prenom', 'sexe', 'fonction',
    'telephone', 'email', 'date_naissance', 'date_prise_service',
]


def _normalize_header(value: str) -> str:
    raw = (value or '').strip().lower()
    spaced = re.sub(r'[\s_]+', ' ', raw)
    underscored = spaced.replace(' ', '_')
    return (
        HEADER_ALIASES.get(raw)
        or HEADER_ALIASES.get(spaced)
        or HEADER_ALIASES.get(underscored)
        or underscored
    )


def _cell_str(value: Any) -> str:
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _parse_date(value: Any) -> date | None:
    if value in (None, ''):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = _cell_str(value)
    if not raw:
        return None
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%d.%m.%Y'):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    raise ValueError(f'Date invalide « {raw} » (AAAA-MM-JJ ou JJ/MM/AAAA)')


def _parse_sexe(value: Any) -> str:
    raw = _cell_str(value).upper()
    if raw in ('M', 'H', 'MASCULIN', 'HOMME', 'GARCON', 'GARÇON'):
        return PersonnelEcole.Sexe.MASCULIN
    if raw in ('F', 'FEMININ', 'FÉMININ', 'FEMME', 'FILLE'):
        return PersonnelEcole.Sexe.FEMININ
    raise ValueError(f'Sexe invalide « {value} » (M ou F)')


def _parse_fonction(value: Any) -> str:
    raw = _cell_str(value).lower()
    if not raw:
        return PersonnelEcole.Fonction.ENSEIGNANT
    key = re.sub(r'\s+', ' ', raw.replace('_', ' ')).strip()
    compact = key.replace(' ', '_')
    mapped = FONCTION_ALIASES.get(key) or FONCTION_ALIASES.get(compact) or FONCTION_ALIASES.get(raw)
    if mapped:
        return mapped
    # valeur déjà code valide
    valid = {c.value for c in PersonnelEcole.Fonction}
    if compact in valid:
        return compact
    raise ValueError(
        f'Fonction invalide « {value} ». '
        f'Valeurs : {", ".join(sorted(valid))}'
    )


def _rows_from_xlsx(data: bytes) -> list[dict[str, str]]:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers_raw = next(rows_iter)
    except StopIteration as exc:
        raise ValueError('Fichier Excel vide.') from exc

    headers = [_normalize_header(_cell_str(h)) for h in headers_raw]
    if not any(headers):
        raise ValueError('En-tête Excel introuvable.')

    missing = [f for f in REQUIRED if f not in headers]
    if missing:
        raise ValueError(
            'Colonnes obligatoires manquantes : ' + ', '.join(missing)
            + '. Attendu : nom, prenom, sexe (+ matricule, fonction, …)'
        )

    rows: list[dict[str, str]] = []
    for i, values in enumerate(rows_iter, start=2):
        if values is None or not any(v not in (None, '') for v in values):
            continue
        row: dict[str, str] = {'_ligne': str(i)}
        for idx, header in enumerate(headers):
            if not header:
                continue
            row[header] = _cell_str(values[idx] if idx < len(values) else '')
        rows.append(row)
    return rows


def _rows_from_csv(data: bytes) -> list[dict[str, str]]:
    for encoding in ('utf-8-sig', 'utf-8', 'cp1252', 'latin-1'):
        try:
            text = data.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = data.decode('utf-8', errors='replace')

    text = text.strip()
    if not text:
        raise ValueError('Fichier vide.')

    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=';,\t|')
    except csv.Error:
        dialect = csv.excel
        dialect.delimiter = ';' if sample.count(';') >= sample.count(',') else ','

    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    if not reader.fieldnames:
        raise ValueError('En-tête CSV introuvable.')

    mapping = {_normalize_header(h): h for h in reader.fieldnames if h}
    missing = [f for f in REQUIRED if f not in mapping]
    if missing:
        raise ValueError('Colonnes obligatoires manquantes : ' + ', '.join(missing))

    rows: list[dict[str, str]] = []
    for i, raw in enumerate(reader, start=2):
        if not any((v or '').strip() for v in raw.values()):
            continue
        row = {canon: (raw.get(original) or '').strip() for canon, original in mapping.items()}
        row['_ligne'] = str(i)
        rows.append(row)
    return rows


def lire_lignes(contenu: bytes, filename: str = '') -> list[dict[str, str]]:
    name = (filename or '').lower()
    if name.endswith(('.xlsx', '.xlsm')) or (
        contenu[:2] == b'PK'  # zip/xlsx
    ):
        try:
            return _rows_from_xlsx(contenu)
        except Exception as exc:  # noqa: BLE001
            if name.endswith(('.xlsx', '.xlsm')):
                raise ValueError(f'Lecture Excel impossible : {exc}') from exc
    return _rows_from_csv(contenu)


def importer_personnel(
    contenu: bytes,
    *,
    ecole_id: int,
    filename: str = '',
    update_existing: bool = True,
) -> dict[str, Any]:
    ecole = Ecole.objects.filter(pk=ecole_id).first()
    if not ecole:
        raise ValueError(f'École introuvable (id={ecole_id}).')

    rows = lire_lignes(contenu, filename=filename)
    created = updated = skipped = 0
    errors: list[dict[str, str]] = []

    with transaction.atomic():
        for row in rows:
            ligne = row.get('_ligne', '?')
            try:
                nom = row.get('nom', '').strip()
                prenom = row.get('prenom', '').strip()
                if not nom or not prenom:
                    raise ValueError('Nom et prénom obligatoires')

                defaults = {
                    'nom': nom,
                    'postnom': row.get('postnom', '').strip(),
                    'prenom': prenom,
                    'sexe': _parse_sexe(row.get('sexe')),
                    'fonction': _parse_fonction(row.get('fonction')),
                    'telephone': row.get('telephone', '').strip()[:20],
                    'email': row.get('email', '').strip(),
                    'date_naissance': _parse_date(row.get('date_naissance')),
                    'date_prise_service': _parse_date(row.get('date_prise_service')),
                    'actif': True,
                }
                matricule = row.get('matricule', '').strip()

                existing = None
                if matricule:
                    existing = PersonnelEcole.objects.filter(
                        ecole=ecole, matricule__iexact=matricule,
                    ).first()
                if not existing:
                    existing = PersonnelEcole.objects.filter(
                        ecole=ecole,
                        nom__iexact=nom,
                        prenom__iexact=prenom,
                        postnom__iexact=defaults['postnom'],
                    ).first()

                if existing:
                    if not update_existing:
                        skipped += 1
                        continue
                    for k, v in defaults.items():
                        setattr(existing, k, v)
                    if matricule:
                        existing.matricule = matricule
                    existing.save()
                    updated += 1
                else:
                    PersonnelEcole.objects.create(
                        ecole=ecole,
                        matricule=matricule,
                        **defaults,
                    )
                    created += 1
            except Exception as exc:  # noqa: BLE001
                errors.append({'ligne': ligne, 'message': str(exc)})

    return {
        'created': created,
        'updated': updated,
        'skipped': skipped,
        'errors': errors[:50],
        'errors_count': len(errors),
        'total': len(rows),
        'ecole_id': ecole.id,
        'ecole_nom': ecole.nom,
    }


def generer_modele_xlsx() -> bytes:
    wb = Workbook()
    info = wb.active
    info.title = 'Instructions'
    info.append(['Champ', 'Description'])
    info.append(['matricule', 'Optionnel — identifiant interne'])
    info.append(['nom / prenom', 'Obligatoires'])
    info.append(['postnom', 'Optionnel'])
    info.append(['sexe', 'Obligatoire — M ou F'])
    info.append([
        'fonction',
        'directeur | directeur_etudes | enseignant | secretaire | '
        'comptable | surveillant | prefet | autre',
    ])
    info.append(['telephone / email', 'Optionnels'])
    info.append(['date_naissance / date_prise_service', 'Optionnels — AAAA-MM-JJ'])
    info.append(['', ''])
    info.append([
        'Note',
        'Ceci importe le personnel identifié (fiche école), '
        'pas les comptes utilisateurs plateforme.',
    ])

    ws = wb.create_sheet('Personnel')
    ws.append(MODELE_HEADERS)
    ws.append([
        'PERS-001', 'KABONGO', 'MUTOMBO', 'Jean', 'M', 'enseignant',
        '+243810000001', 'jean.kabongo@exemple.cd', '1985-03-12', '2015-09-01',
    ])
    ws.append([
        'PERS-002', 'MUKENDI', '', 'Marie', 'F', 'directeur',
        '+243810000002', 'marie.mukendi@exemple.cd', '1978-07-21', '2010-01-15',
    ])
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max(len(str(c.value or '')) for c in col)
            sheet.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 14), 48)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def reponse_modele_xlsx() -> HttpResponse:
    content = generer_modele_xlsx()
    response = HttpResponse(
        content,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="modele_import_personnel.xlsx"'
    return response
